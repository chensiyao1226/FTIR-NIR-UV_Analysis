import os
import csv
import numpy as np
import pandas as pd
import openpyxl as xl
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import scipy.signal as signal

#============= normalization ===============
def SNV(input_data):
    # Define a new array and populate it with the corrected data
    output_data = np.zeros_like(input_data)
    for i in range(input_data.shape[0]):
        # Apply correction
        output_data[i, :] = (input_data[i, :] - np.mean(input_data[i, :])) / np.std(input_data[i, :])

    return output_data

#================save============
def save(file_root, excel_name, label, data_matrix):
    file_magnet = file_root+"\\"+excel_name+".xlsx"
    try:
        if os.access(file_magnet, os.F_OK):
            print("modifying...")
            book = xl.load_workbook(file_magnet)
            sheet = book[excel_name]

            sheet.append([label])
            for i in range(len(data_matrix)):
                sheet.append(data_matrix[i])
        else:
            print("creating...")
            book = xl.Workbook()
            sheet = book.active
            sheet.title = excel_name

            sheet.append([label])
            for i in range(len(data_matrix)):
                sheet.append(data_matrix[i])

        book.save(file_magnet)
        print("Saved")
    except:
        print("Falied!!! Close the file")


#===============arrange sequence===============
def takeSecond(elem):
    return elem[1]
def arrange(data_matrix, reverse_F):
    data_matrix.sort(key=takeSecond, reverse=reverse_F)
    print(data_matrix)



path_root = input('input folder root:')
rep_t = 3
#============collect data=================
counter = 0
content_list=np.zeros((1869,2))
content_matrix=[]

for root, dirs, files in os.walk(path_root):
    for filename in files:
        if ('.csv' in filename):
            file_path = os.path.join(root, filename)
            print(filename)
            samp_sheet = pd.read_csv(file_path, skiprows=19, encoding = "shift-jis", header=None)
            samp_dat = samp_sheet.iloc[0:1869, 0:2]
            content_list = samp_dat.values




            if(counter==0):
                content_list[:, 0] = [float(x) for x in content_list[:, 0]]
                content_matrix.append(content_list[:,0])

            content_list[:, 1] = [float(x) for x in content_list[:, 1]]
            content_matrix.append(content_list[:,1])

            counter += 1

#=========cut================
Edge1L = 880; Edge1Lnum=100000 ##minEdge
Edge1H = 1900; Edge1Hnum=1 ##minEdge
Edge2L = 2400; Edge2Lnum=100000 ##maxEdge
Edge2H = 3820; Edge2Hnum=1 ##maxEdge


print(len(content_matrix[0]))

for i in range (len(content_matrix[0])):
    if((content_matrix[0][i]>Edge1L) and (i<Edge1Lnum)):
        Edge1Lnum = i
    if ((content_matrix[0][i] < Edge1H) and (i > Edge1Hnum)):
        Edge1Hnum = i
    if ((content_matrix[0][i] > Edge2L) and (i < Edge2Lnum)):
        Edge2Lnum = i
    if ((content_matrix[0][i] < Edge2H) and (i > Edge2Hnum)):
        Edge2Hnum = i




content_cutted=[x[Edge1Lnum:Edge2Hnum] for x in content_matrix]
content_cutted = np.array(content_cutted)

norm_cutted_data = content_cutted
norm_cutted_data[1:] = SNV(content_cutted[1:])

sum_cutted_data=np.zeros((int(counter/rep_t),norm_cutted_data.shape[1]))

for i in range(int(counter/rep_t)):
    for k in range(rep_t):
         sum_cutted_data[i] = sum_cutted_data[i]+norm_cutted_data[3*i+1+k,:]

sum_cutted_data = sum_cutted_data/rep_t

plt.figure()
for i in range(int(counter/rep_t)):
    lab = i + 1
    if (lab > 4):
        lab+=2
    plt.plot(norm_cutted_data[0, :], sum_cutted_data[i], label=lab, alpha=0.6)
plt.legend()
plt.xlabel('Wave Number (cm-1)')
plt.ylabel('Normalized Absorbance')
plt.gca().invert_xaxis()


#=============== ogirinal data find peak =========
peakOH = []
peakOL = []
indicesOH = signal.find_peaks(sum_cutted_data[0], height=0.15, distance=10, prominence=0.1)
indicesOL = signal.find_peaks(-sum_cutted_data[0], height=0.15, distance=10, prominence=0.1)

for k in range(len(indicesOH[0])):
    nDH = indicesOH[0][k]
    peakOH.append((round(norm_cutted_data[0][nDH]), sum_cutted_data[i][nDH]))
    #plt.plot(norm_cutted_data[0][nDH], sum_cutted_data[i][nDH], "x")
    strH = "Dif Peak:" + str(norm_cutted_data[0][nDH]) + ":" + str(sum_cutted_data[i][nDH])
    print(strH)
for k in range(len(indicesOL[0])):
    nDL = indicesOL[0][k]
    peakOL.append((round(norm_cutted_data[0][nDL]), sum_cutted_data[i][nDL]))
    #plt.plot(norm_cutted_data[0][nDL], sum_cutted_data[i][nDL], "x")
    strL = "Dif Peak:" + str(norm_cutted_data[0][nDL]) + ":" + str(sum_cutted_data[i][nDL])
    print(strL)
plt.show()

arrange(peakOH, True)
arrange(peakOL, False)
labelO = "original" + str(i)
save(path_root, "peak record original FTIR", labelO, peakOH)
save(path_root, "peak record original FTIR", labelO, peakOL)

plt.show()




#==============Savitzky-Golay=============
from scipy.signal import savgol_filter

savgol=content_cutted
savgol[0] = content_cutted[0]
for i in range(1,content_cutted.shape[0]):
    savgol[i]=savgol_filter(content_cutted[i], window_length = 15, polyorder = 2, deriv = 2)
    savgol[i,0:35]=0
    savgol[i,(len(savgol[0])-35):(len(savgol[0]))] = 0
    #savgol[i, (Edge1Hnum-Edge1Lnum): (Edge2Lnum-Edge1Lnum)]=0


plt.figure()
plt.plot(savgol[0], savgol[1])
plt.show()



norm_data = savgol
norm_data[1:] = savgol[1:]######################gai


'''
for i in range(content_cutted.shape[0]-1):
    plt.figure()
    plt.plot(savgol[0], norm_data[i])
    plt.show()
'''

#===========PCA===========

from mpl_toolkits.mplot3d import Axes3D
from sklearn.decomposition import PCA
pca = PCA(int(counter/rep_t))
pca.fit(norm_data[1:])

# データを主成分空間に写像
features = pca.transform(norm_data[1:])

# 主成分得点
point = pd.DataFrame(features, columns=["PC{}".format(x + 1) for x in range(int(((norm_data.shape[0])-1)/rep_t))])
print(point)


# 第一主成分と第二主成分でプロットする
# np.random.rand(n)产生1*n数组，元素大小0-1
from mpl_toolkits.mplot3d import Axes3D
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')
ax.scatter(features[:, 0], features[:, 1], features[:, 2], c=list(norm_data[1:, 0]))

ax.set_xlabel('PC1')
ax.set_ylabel('PC2')
ax.set_zlabel('PC3')


for i in range(len(features[:, 0])):
    ax.text(features[i, 0],features[i, 1], features[i, 2], i)

plt.show()

'''
plt.figure(figsize=(6, 6))
plt.scatter(features[:, 0], features[:, 1], alpha=0.8, c=list(norm_data[1:, 0]))
plt.grid()
plt.xlabel("PC1")
plt.ylabel("PC2")
'''


# 累積寄与率を図示する
import matplotlib.ticker as ticker
plt.gca().get_xaxis().set_major_locator(ticker.MaxNLocator(integer=True))
plt.plot([0] + list( np.cumsum(pca.explained_variance_ratio_)), "-o")
plt.xlabel("Number of principal components")
plt.ylabel("Cumulative contribution rate")
plt.grid()
plt.show()


# PCA の固有ベクトル
Intr_matrix = pd.DataFrame(pca.components_, columns=norm_data[0,:], index=["PC{}".format(x + 1) for x in range(int(((norm_data.shape[0]-1))/rep_t))])
print(Intr_matrix)



#================plot Loading===================
princ_amount = int(input('principle amount:'))
for i in range(princ_amount):
    plt.figure()
    plt.plot(norm_data[0, :], Intr_matrix.values[i], color='red', alpha=0.4)

    indicesVH = signal.find_peaks(Intr_matrix.values[i], height=0.05)
    indicesVL = signal.find_peaks(-Intr_matrix.values[i], height=0.02)

    for k in range(len(indicesVH[0])):
        nVH = indicesVH[0][k]
        plt.plot(norm_data[0][nVH], Intr_matrix.values[i][nVH], "x")
        strH = str(norm_data[0][nVH]) + ":" + str(Intr_matrix.values[i][nVH])
        print(strH)
    for k in range(len(indicesVL[0])):
        nVL = indicesVL[0][k]
        plt.plot(norm_data[0][nVL], Intr_matrix.values[i][nVL], "x")
        strL = str(norm_data[0][nVL]) + ":" + str(Intr_matrix.values[i][nVL])
        print(strL)

    plt.show()
'''
from pandas import plotting
plotting.scatter_matrix(pd.DataFrame(features,
                        columns=["PC{}".format(x + 1) for x in range(norm_data.shape[0]-1)]),
                        figsize=(8, 8), c=list(norm_data[1:, 0]), alpha=0.5)
plt.show()
'''


#============ transfer ================
print("transfer")
t1=np.zeros((int(counter/rep_t),norm_data.shape[1]))
t2=np.zeros((int(counter/rep_t),norm_data.shape[1]))

t1peak=np.zeros((int(counter/rep_t),norm_data.shape[1]))

for i in range(int(counter/rep_t)):
    for k in range(rep_t):
        t1[i] = t1[i] + norm_data[3*i+1+k,:]*Intr_matrix.values[0]
        t2[i] = t2[i] + norm_data[3 * i + 1 + k, :] * Intr_matrix.values[1]

'''
    plt.figure()
    plt.plot(norm_data[0,:], t1[i], color='red', alpha=0.4)
    plt.plot(norm_data[0, :], t2[i], color='green', alpha=0.4)

    # =========== Extreme point ============
    indicesH = signal.find_peaks(t1[i], height=0.1)
    indicesL = signal.find_peaks(-t1[i], height=0.1)

    for k in range(len(indicesH[0])):
        nH=indicesH[0][k]
        plt.plot(norm_data[0][nH], t1[i][nH], "x")
        strH = str(norm_data[0][nH]) + ":" + str(Intr_matrix.values[0][nH])
        print(strH)
    for k in range(len(indicesL[0])):
        nL = indicesL[0][k]
        plt.plot(norm_data[0][nL], t1[i][nL], "x")
        strL = str(norm_data[0][nL]) + ":" + str(Intr_matrix.values[0][nL])
        print(strL)

    plt.show()
'''






ntime=3

PC=np.zeros(len(Intr_matrix.values[0]))

for i in range(princ_amount):
    PC=PC+Intr_matrix.values[i]


#============= difference ===============
sum_data=np.zeros((int(counter/rep_t),norm_data.shape[1]))

for i in range(int(counter/rep_t)):
    for k in range(rep_t):
         sum_data[i] = sum_data[i]+norm_data[3*i+1+k,:]

sum_data = sum_data/rep_t

plt.figure()
for i in range(int(counter/rep_t)):
    lab = i + 1
    if (lab > 4):
        lab += 2
    plt.plot(norm_data[0, :], sum_data[i], label=lab, alpha=0.4)
plt.legend()
plt.xlabel('Wave Number (cm-1)')
plt.ylabel('Normalized 2nd Derivation')
plt.show()

print("difference")
ref = int(input('input reference number:'))

ref_data = sum_data[ref]
dif_data=np.zeros((int(counter/rep_t),norm_data.shape[1]))


for i in range(int(counter/rep_t)):
    dif_data[i]= sum_data[i] - ref_data


for i in range(int(counter/rep_t)):
    peakPH = []
    peakPL = []
    peakDH = []
    peakDL = []

    plt.figure()
    plt.plot(norm_data[0, :], dif_data[i], color='green', alpha=0.6)
    plt.plot(norm_data[0, :], ntime*PC, color='red', alpha=0.3)

    # =========== Extreme point ============
    indicesPH = signal.find_peaks(ntime*PC, height=0.1)
    indicesPL = signal.find_peaks(-ntime*PC, height=0.1)

    for k in range(len(indicesPH[0])):
        nPH = indicesPH[0][k]
        peakPH.append((round(norm_data[0][nPH]), ntime*PC[nPH]))
        plt.plot(norm_data[0][nPH], ntime*PC[nPH], "x")
        strH = str(norm_data[0][nPH]) + ":" + str(ntime*PC[nPH])
        print(strH)
    for k in range(len(indicesPL[0])):
        nPL = indicesPL[0][k]
        peakPL.append((round(norm_data[0][nPL]), ntime * PC[nPL]))
        plt.plot(norm_data[0][nPL], ntime*PC[nPL], "x")
        strL = str(norm_data[0][nPL]) + ":" + str(ntime*PC[nPL])
        print(strL)

    arrange(peakPH, True)
    arrange(peakPL, False)
    labelP = "PCALL"+str(i)
    save(path_root, "peak record FTIR", labelP, peakPH)
    save(path_root, "peak record FTIR", labelP, peakPL)

    #================derivation=============
    indicesDH = signal.find_peaks(dif_data[i], height=0.15)
    indicesDL = signal.find_peaks(-dif_data[i], height=0.15)

    for k in range(len(indicesDH[0])):
        nDH = indicesDH[0][k]
        peakDH.append((round(norm_data[0][nDH]), dif_data[i][nDH]))
        plt.plot(norm_data[0][nDH], dif_data[i][nDH], "x")
        strH = "Dif Peak:"+str(norm_data[0][nDH]) + ":" + str(dif_data[i][nDH])
        print(strH)
    for k in range(len(indicesDL[0])):
        nDL = indicesDL[0][k]
        peakDL.append((round(norm_data[0][nDL]), dif_data[i][nDL]))
        plt.plot(norm_data[0][nDL], dif_data[i][nDL], "x")
        strL = "Dif Peak:"+str(norm_data[0][nDL]) + ":" + str(dif_data[i][nDL])
        print(strL)
    plt.show()

    arrange(peakDH, True)
    arrange(peakDL, False)
    labelD = "Differentiation"+str(i)
    save(path_root, "peak record FTIR", labelD, peakDH)
    save(path_root, "peak record FTIR", labelD, peakDL)
'''
#=========== >PC3 ===========
R=np.zeros((int(counter/rep_t),norm_data.shape[1]))
for i in range(int(counter/rep_t)):
    R[i] = norm_data[i + 1, :]- t1[i]-t2[i]

    plt.figure()
    plt.plot(norm_data[0, :], R[i], color='red', alpha=0.6)
    plt.plot(norm_data[0, :], norm_data[i + 1, :], color='green', alpha=0.2)
    plt.show()
'''



#=========save==============
'''
try:
    exc_save_magn = path_root + '\\' + 'summary.csv'

    book = xl.Workbook()
    sheet1 = book.create_sheet("original")

    for i in range (counter+1):
        list = content_matrix[i]
        list = list.tolist()
        sheet1.append(list)

    book.save(exc_save_magn)
    print("Saved")

except:
    print("Falied!!!")
'''






